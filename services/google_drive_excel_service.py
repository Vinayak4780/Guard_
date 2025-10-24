"""
Simple Google Drive Excel Service - No API Key Required
Uses local Excel files that can be manually synced to Google Drive
"""

import os
import logging
import asyncio
from datetime import datetime
from typing import Optional, Dict, Any, List
from io import BytesIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import settings

logger = logging.getLogger(__name__)

class GoogleDriveExcelService:
    """Simple Excel service that creates files for Google Drive sync"""
    
    def __init__(self):
        # Create a local Excel folder that can be synced to Google Drive
        self.excel_folder = "./google_drive_excel"
        self.excel_file_name = settings.EXCEL_FILE_NAME or "guard_scan_reports.xlsx"
        self.update_interval = getattr(settings, 'UPDATE_INTERVAL_SECONDS', 1)
        
        # Queue for batch updates
        self.update_queue = []
        self.queue_lock = asyncio.Lock()
        
        # Excel headers
        self.headers = [
            "Timestamp", "Date", "Time", "Guard Name", "Guard Email", "Employee Code",
            "Supervisor Name", "Area City", "QR Location", "Latitude", "Longitude",
            "Distance (meters)", "Status", "Address", "Landmark", "Remarks"
        ]
        
        # Initialize service
        self._initialize_local_service()
        logger.info(f"📊 Simple Excel service initialized - Folder: {self.excel_folder}")
    
    def _initialize_local_service(self) -> bool:
        """Initialize local Excel folder"""
        try:
            # Create Excel folder if it doesn't exist
            os.makedirs(self.excel_folder, exist_ok=True)
            
            # Create a README file with instructions
            readme_path = os.path.join(self.excel_folder, "README.txt")
            with open(readme_path, 'w') as f:
                f.write("""
GOOGLE DRIVE SYNC INSTRUCTIONS

1. Install Google Drive Desktop App:
   - Download from: https://www.google.com/drive/download/
   - Install and sign in with your Google account

2. Sync this folder to Google Drive:
   - Copy this entire folder: google_drive_excel
   - Paste it into your Google Drive folder on your computer
   - It will automatically sync to Google Drive online

3. All Excel files will be automatically created here and synced to Google Drive!

FOLDER CONTENTS:
- Excel files are created per supervisor area
- Files are updated in real-time as guards scan QR codes
- No API keys needed - just Google Drive desktop sync!
""")
            
            logger.info("✅ Local Excel service initialized successfully")
            logger.info(f"📁 Excel files will be saved to: {os.path.abspath(self.excel_folder)}")
            logger.info("💡 Install Google Drive desktop app to auto-sync to Google Drive")
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to initialize local service: {e}")
            return False
    
    async def add_scan_to_queue(self, scan_data: Dict[str, Any]) -> bool:
        """Add scan data to the update queue for batch processing"""
        try:
            async with self.queue_lock:
                self.update_queue.append(scan_data)
                logger.debug(f"📝 Added scan to queue. Queue size: {len(self.update_queue)}")
                return True
                
        except Exception as e:
            logger.error(f"❌ Failed to add scan to queue: {e}")
            return False
    
    async def process_update_queue(self) -> bool:
        """Process all queued updates and update Excel files"""
        if not self.update_queue:
            return True
        
        try:
            async with self.queue_lock:
                # Get all pending updates
                pending_updates = self.update_queue.copy()
                self.update_queue.clear()
            
            if not pending_updates:
                return True
            
            # Process updates by supervisor area
            success = self._process_scans_by_area(pending_updates)
            
            if success:
                logger.info(f"✅ Processed {len(pending_updates)} scan updates to Excel files")
            else:
                # Re-add failed updates to queue
                async with self.queue_lock:
                    self.update_queue.extend(pending_updates)
                logger.error(f"❌ Failed to process updates, re-queued {len(pending_updates)} items")
            
            return success
            
        except Exception as e:
            logger.error(f"❌ Error processing update queue: {e}")
            return False
    
    def _process_scans_by_area(self, scan_updates: List[Dict[str, Any]]) -> bool:
        """Process scan updates and save to appropriate Excel files"""
        try:
            # Group scans by supervisor area
            scans_by_area = {}
            for scan_data in scan_updates:
                area = scan_data.get('area_city', 'Unknown_Area')
                if area not in scans_by_area:
                    scans_by_area[area] = []
                scans_by_area[area].append(scan_data)
            
            # Process each area
            for area, scans in scans_by_area.items():
                success = self._update_area_excel_file(area, scans)
                if not success:
                    logger.error(f"❌ Failed to update Excel for area: {area}")
                    return False
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to process scans by area: {e}")
            return False
    
    def _update_area_excel_file(self, area: str, scans: List[Dict[str, Any]]) -> bool:
        """Update Excel file for specific area"""
        try:
            # Create safe filename
            safe_area = "".join(c for c in area if c.isalnum() or c in (' ', '-', '_')).rstrip()
            excel_file_path = os.path.join(self.excel_folder, f"{safe_area}_Guard_Scans.xlsx")
            
            # Load or create workbook
            if os.path.exists(excel_file_path):
                wb = load_workbook(excel_file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = f"{area} Scans"
                
                # Add headers
                ws.append(self.headers)
                self._format_header_row(ws)
                
                # Add title row
                ws.insert_rows(1)
                ws['A1'] = f"Guard Scan Report - {area}"
                from utils.timezone_utils import get_current_ist_string
                ws['B1'] = get_current_ist_string()
                self._format_title_row(ws)
            
            # Add scan data
            for scan_data in scans:
                row_data = [
                    scan_data.get('timestamp', ''),
                    scan_data.get('date', ''),
                    scan_data.get('time', ''),
                    scan_data.get('guard_name', ''),
                    scan_data.get('guard_email', ''),
                    scan_data.get('employee_code', ''),
                    scan_data.get('supervisor_name', ''),
                    scan_data.get('area_city', ''),
                    scan_data.get('qr_location', ''),
                    scan_data.get('latitude', ''),
                    scan_data.get('longitude', ''),
                    scan_data.get('distance_meters', ''),
                    scan_data.get('status', ''),
                    scan_data.get('address', ''),
                    scan_data.get('landmark', ''),
                    scan_data.get('remarks', '')
                ]
                
                ws.append(row_data)
            
            # Save file
            wb.save(excel_file_path)
            logger.debug(f"✅ Updated Excel file: {excel_file_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to update area Excel file: {e}")
            return False
    
    def _format_header_row(self, worksheet):
        """Format the header row with styling"""
        try:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Header is in row 2 (after title)
            for cell in worksheet[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
        except Exception as e:
            logger.warning(f"⚠️ Failed to format header row: {e}")
    
    def _format_title_row(self, worksheet):
        """Format the title row"""
        try:
            title_font = Font(bold=True, size=14)
            title_alignment = Alignment(horizontal="left", vertical="center")
            
            worksheet['A1'].font = title_font
            worksheet['A1'].alignment = title_alignment
            worksheet['B1'].font = Font(size=12)
            worksheet['B1'].alignment = title_alignment
            
        except Exception as e:
            logger.warning(f"⚠️ Failed to format title row: {e}")
    
    async def start_background_updates(self):
        """Start background task for processing queued updates"""
        logger.info(f"🔄 Starting background updates every {self.update_interval} second(s)")
        
        while True:
            try:
                await self.process_update_queue()
                await asyncio.sleep(self.update_interval)
                
            except asyncio.CancelledError:
                logger.info("🛑 Background update task cancelled")
                break
            except Exception as e:
                logger.error(f"❌ Error in background update task: {e}")
                await asyncio.sleep(5)  # Wait before retrying

# Create global instance
google_drive_excel_service = GoogleDriveExcelService()
